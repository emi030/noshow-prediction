import re
import pandas as pd
import numpy as np
from datetime import datetime

# ── Paste your daily schedule here ──
# Copy and paste directly from your EHR daily schedule view
SCHEDULE_TEXT = """
# Paste your clinic's daily schedule here
# Example format:
# * 9:00 AM30min
# Patient Name
# MM-DD-YYYY(phone number)
# Appointment Type
# Insurance Name
"""

# ── Insurance Type Mapping ──
def classify_insurance(ins):
    if pd.isna(ins) or str(ins).strip() == '':
        return 'Personal Payment (Cash - No Insurance)'
    # Remove "Eligibility Issue" prefix — does not mean no insurance
    ins = re.sub(r'eligibility\s*issue', '', ins.lower(), flags=re.IGNORECASE).strip()
    if any(x in ins for x in ['self-pay', 'cash', 'self pay']):
        return 'Personal Payment (Cash - No Insurance)'
    if 'medicare' in ins:
        return 'Medicare Part B'
    if 'medicaid' in ins:
        return 'Medicaid'
    if 'hmo' in ins:
        return 'Health Maintenance Organization (HMO)'
    if any(x in ins for x in ['bcbs', 'blue cross', 'anthem', 'united', 'aetna', 'cigna', 'humana', 'premera']):
        return 'Commercial'
    return 'Other'

# ── Appointment Type Mapping ──
def classify_appttype(appt):
    if pd.isna(appt) or str(appt).strip() == '':
        return 'Other'
    appt = appt.lower()
    if any(x in appt for x in ['follow up', 'follow-up', 'followup', 'f/u', 'lab']):
        return 'Follow Up'
    if any(x in appt for x in ['ultrasound', 'us ', 'doppler', 'fibroscan']):
        return 'Ultrasound Testing'
    if any(x in appt for x in ['wellness', 'annual', 'well visit']):
        return 'Wellness Visit'
    if any(x in appt for x in ['new patient', 'new visit', 'immigration']):
        return 'New Patient'
    if 'weight' in appt:
        return 'Weight Management'
    if any(x in appt for x in ['virtual', 'telehealth', 'telemedicine', 'privia']):
        return 'Privia Virtual Visit'
    if 'established' in appt:
        return 'Established Patient'
    return 'Other'

# ── Calculate Age ──
def calc_age(dob_str):
    try:
        dob = datetime.strptime(dob_str, '%m-%d-%Y')
        today = datetime.today()
        return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
    except:
        return np.nan

# ── Parse Schedule Text ──
def parse_schedule(text):
    patients = []
    raw = text.split('\n')
    raw = [l.replace('\u200b','').replace('\u00a0','').replace('\r','').strip() for l in raw]
    raw = [l for l in raw if l and not l.startswith('#')]

    # Rejoin lines that got split (30min, phone fragments)
    joined = []
    cur = ''
    for line in raw:
        if (re.match(r'^\d{2}min$', line) or re.match(r'^\(\d{3}\)', line)) and cur:
            cur += line
        else:
            if cur:
                joined.append(cur)
            cur = line
    if cur:
        joined.append(cur)

    time_re = re.compile(r'(\d{1,2}:\d{2}\s*[AP]M)', re.IGNORECASE)
    dob_re  = re.compile(r'^(\d{2}-\d{2}-\d{4})')

    def is_skip(line):
        if re.match(r'^(open|lunch)$', line, re.IGNORECASE): return True
        if re.search(r'pmg_|office\*', line, re.IGNORECASE): return True
        if re.match(r'^[A-Za-z\s,\.\']+,?\s*(MD|DO|NP|PA|RN|DPM|PhD|CRNP|FNP|LAc)\.?\s*$', line, re.IGNORECASE): return True
        if re.match(r'^(jcm\s|jcm$)', line, re.IGNORECASE): return True
        if re.match(r'^eligibility\s*issue$', line, re.IGNORECASE): return True
        return False

    i = 0
    while i < len(joined):
        line = joined[i]
        if is_skip(line): i += 1; continue

        tm = time_re.search(line)
        if tm:
            time = tm.group(1).replace(' ', '')
            i += 1

            if i < len(joined) and re.match(r'^(open|lunch)$', joined[i], re.IGNORECASE):
                i += 1; continue
            if i >= len(joined): continue
            if joined[i].startswith('*') and not time_re.search(joined[i]): continue

            name = joined[i]
            if not name or len(name) < 2 or name[0].isdigit(): continue
            i += 1

            dob = None
            appt_desc = None
            insurance = None

            if i < len(joined) and dob_re.match(joined[i]):
                dob = dob_re.match(joined[i]).group(1)
                i += 1

            if i < len(joined) and not joined[i].startswith('*') and not dob_re.match(joined[i]) and not re.match(r'^(open|lunch)$', joined[i], re.IGNORECASE):
                appt_desc = joined[i]
                i += 1

            if i < len(joined) and not joined[i].startswith('*') and not dob_re.match(joined[i]) and not re.match(r'^(open|lunch)$', joined[i], re.IGNORECASE):
                insurance = joined[i]
                i += 1

            patients.append({
                'time':          time,
                'name':          name.strip(),
                'dob':           dob,
                'appt_desc':     appt_desc,
                'insurance_raw': insurance
            })
            continue
        i += 1

    return patients

# ── Risk Weights (from historical analysis) ──
INSURANCE_RISK = {
    'Personal Payment (Cash - No Insurance)': 0.77,
    'Medicaid':                               0.10,
    'Other':                                  0.08,
    'Commercial':                             0.07,
    'Group Policy':                           0.07,
    'Health Maintenance Organization (HMO)':  0.05,
    'Medicare Part B':                        0.04,
    'Supplemental Policy':                    0.01,
}

def age_risk(age):
    if pd.isna(age):  return 0.20
    if age < 30:      return 0.28
    if age < 40:      return 0.22
    if age < 50:      return 0.18
    if age < 60:      return 0.14
    return 0.08

def risk_label(score):
    if score >= 0.45: return 'HIGH RISK ⚠️ '
    if score >= 0.20: return 'MEDIUM    🟡'
    return            'LOW       ✅'

# ── Main ──
patients = parse_schedule(SCHEDULE_TEXT)

if not patients:
    print("No patients found. Please paste your daily schedule into SCHEDULE_TEXT.")
else:
    results = []
    for p in patients:
        age       = calc_age(p['dob']) if p['dob'] else np.nan
        ins_type  = classify_insurance(p['insurance_raw'])
        appt_type = classify_appttype(p['appt_desc'])
        ins_score = INSURANCE_RISK.get(ins_type, 0.07)
        a_score   = age_risk(age)
        score     = round((ins_score * 0.55) + (a_score * 0.45), 3)

        results.append({
            'Risk':      risk_label(score),
            'Score':     score,
            'Time':      p['time'],
            'Name':      p['name'],
            'Age':       int(age) if not pd.isna(age) else 'N/A',
            'Appt Type': appt_type,
            'Insurance': ins_type,
        })

    df = pd.DataFrame(results).sort_values('Score', ascending=False)

    print("\n" + "="*85)
    print("  DAILY NO-SHOW RISK LIST")
    print("  Generated:", datetime.today().strftime('%Y-%m-%d'))
    print("="*85)
    print(f"{'Risk':<16} {'Time':<12} {'Name':<22} {'Age':<5} {'Insurance':<35} {'Appt Type'}")
    print("-"*85)
    for _, row in df.iterrows():
        print(f"{row['Risk']:<16} {row['Time']:<12} {row['Name']:<22} {str(row['Age']):<5} {row['Insurance']:<35} {row['Appt Type']}")
    print("="*85)
    print(f"\nTotal: {len(df)}  |  High: {len(df[df['Risk'].str.contains('HIGH')])}  |  Medium: {len(df[df['Risk'].str.contains('MEDIUM')])}  |  Low: {len(df[df['Risk'].str.contains('LOW')])}")

    # ── Save to Excel ──
    from openpyxl.styles import PatternFill
    output_file = f"noshow_risk_{datetime.today().strftime('%Y%m%d')}.xlsx"
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Risk List')
        ws = writer.sheets['Risk List']
        red    = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        yellow = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        green  = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            risk = str(row[0].value)
            fill = red if 'HIGH' in risk else yellow if 'MEDIUM' in risk else green
            for cell in row:
                cell.fill = fill
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(max_len, 40)

    print(f"Excel saved: {output_file}")
